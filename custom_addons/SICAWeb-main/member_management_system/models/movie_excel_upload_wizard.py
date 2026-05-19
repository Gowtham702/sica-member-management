from odoo import models, fields, api, _
import base64
from odoo.exceptions import ValidationError
import io
import xlrd
from openpyxl import load_workbook
import xlsxwriter
import datetime

class MovieExcelUploadWizard(models.TransientModel):
    _name = 'movie.excel.upload.wizard'
    _description = 'Upload Movie Excel Wizard'

    excel_file = fields.Binary('Excel File', required=False)
    file_name = fields.Char('Filename')

    def action_import(self):
        if not self.excel_file:
            raise ValidationError("Please upload a valid Excel file.")

        try:
            decoded_file = base64.b64decode(self.excel_file)
            workbook = load_workbook(filename=io.BytesIO(decoded_file), data_only=True)
        except Exception as e:
            raise ValidationError(f"Invalid or corrupted Excel file: {e}")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    sno = row[0]
                    date_val = row[1]

                    # Take Excel values EXACTLY as they are
                    movie_name = row[2] or ''
                    dop_name = row[3] or ''
                    production = row[4] or ''
                    team_member = row[5] or ''
                    movie_link = row[6] or ''
                    project_type = row[7] or ''
                    channel_name = row[8] or ''
                    member_type = row[9] or ''
                    language = row[10] or ''
                    membership_no = row[11] or ''
                    grade = row[12] or ''
                    release_state = row[13] or ''

                    # --- DATE PARSER (ONLY LOGIC WE KEEP) ---
                    parsed_date = fields.Date.today()

                    if isinstance(date_val, datetime.date):
                        parsed_date = date_val
                    elif isinstance(date_val, str):
                        for fmt in ("%d %B %Y", "%d-%b-%y", "%d/%m/%Y"):
                            try:
                                parsed_date = datetime.datetime.strptime(date_val, fmt).date()
                                break
                            except:
                                continue

                    # Create the movie record EXACTLY as Excel provides
                    self.env['movie.list'].create({
                        'sequence': int(sno) if sno else 0,
                        'date': parsed_date,
                        'movie_name': movie_name,
                        'dop_name': dop_name,
                        'production_companies': production,
                        'team_member': team_member,
                        'movie_link': movie_link,
                        'project_type': project_type,
                        'channel_name': channel_name,
                        'member_type': member_type,
                        'language': language,
                        'membership_no': membership_no,
                        'grade': grade,
                        'release_state': release_state,
                    })

                except Exception as e:
                    raise ValidationError(
                        f"Error in sheet '{sheet_name}' row {row_index}: {e}"
                    )

        return {
            'type': 'ir.actions.act_window',
            'name': 'Movie List',
            'res_model': 'movie.list',
            'view_mode': 'tree,form',
            'target': 'current',
        }

    def action_download_template(self):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet('Movie Template')

        header_format = workbook.add_format({'bold': True, 'bg_color': '#DCE6F1', 'border': 1})
        text_format = workbook.add_format({'border': 1})

        headers = [
            'S.No',
            'Date (e.g. 25 December 2025)',
            'Movie Name',
            'DOP Name',
            'Production Company',
            'Team Member',
            'Movie Link',
            'Project Type (film/series/serial/others)',
            'Channel Name',
            'Member Type',
            'Language',
            'Membership No.',
            'Grade',
            'Release State (Released/Not Released)',
        ]

        for col, header in enumerate(headers):
            sheet.write(0, col, header, header_format)
            sheet.set_column(col, col, 25)

        example_row = [
            1, '25 December 2025', 'Avatar 3', 'John Doe', 'Lightstorm',
            'Team A', 'https://example.com', 'film', 'HBO',
            'member', 'English', 'M12345', 'Active', 'Released'
        ]

        for col, value in enumerate(example_row):
            sheet.write(1, col, value, text_format)

        workbook.close()
        output.seek(0)
        file_data = output.read()

        self.file_name = 'movie_template.xlsx'
        self.excel_file = base64.b64encode(file_data)

        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model={self._name}&id={self.id}&field=excel_file&filename_field=file_name&download=true",
            'target': 'self',
        }



    # def action_import(self):
    #     if not self.excel_file:
    #         raise ValidationError("Please upload a valid Excel file.")

    #     workbook = xlrd.open_workbook(file_contents=base64.b64decode(self.excel_file))
    #     sheet = workbook.sheet_by_index(0)

    #     for row in range(1, sheet.nrows):
    #         sno = sheet.cell(row, 0).value
    #         date_val = sheet.cell(row, 1).value
    #         movie_name = sheet.cell(row, 2).value
    #         dop_name = sheet.cell(row, 3).value
    #         production = sheet.cell(row, 4).value if sheet.ncols > 4 else ''
    #         team_member = sheet.cell(row, 5).value if sheet.ncols > 5 else ''
    #         movie_link = sheet.cell(row, 6).value if sheet.ncols > 6 else ''
    #         if isinstance(date_val, float):
    #             parsed_date = datetime.datetime(*xlrd.xldate_as_tuple(date_val, workbook.datemode)).date()
    #         else:
    #             try:
    #                 parsed_date = datetime.datetime.strptime(date_val, "%d %b %Y").date()
    #             except Exception:
    #                 parsed_date = fields.Date.today()

    #         self.env['movie.list'].create({
    #             'sequence': int(sno),
    #             'date': parsed_date,
    #             'movie_name': movie_name,
    #             'dop_name': dop_name,
    #             'production_companies': production,
    #             'team_member': team_member,
    #             'movie_link': movie_link,
    #         })

    #     # Redirect to movie.list tree  
    #     return {
    #         'type': 'ir.actions.act_window',
    #         'name': 'Movie List',
    #         'res_model': 'movie.list',
    #         'view_mode': 'tree,form',
    #         'target': 'current',
    #     }