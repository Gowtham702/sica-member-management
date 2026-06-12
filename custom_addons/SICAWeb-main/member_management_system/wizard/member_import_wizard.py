# -*- coding: utf-8 -*-

from odoo import models, fields, _
from odoo.exceptions import UserError
import base64
import io
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    openpyxl = None


class MemberImportWizard(models.TransientModel):
    _name = 'member.import.wizard'
    _description = 'Member Import Wizard'

    paid_till_file = fields.Binary(string='Excel File (Paid Till)')
    paid_till_file_name = fields.Char(string='File Name')

    sica_file = fields.Binary(string='Excel File (SICA/CBT)')
    sica_file_name = fields.Char(string='File Name')

    expense_file = fields.Binary(string='Excel File (Expense Receipt)')
    expense_file_name = fields.Char(string='File Name')

    total_records = fields.Integer(string='Total Records', readonly=True)
    updated_count = fields.Integer(string='Updated / Inserted', readonly=True)
    not_found_count = fields.Integer(string='Not Found', readonly=True)
    error_count = fields.Integer(string='Errors', readonly=True)
    skipped_count = fields.Integer(string='Skipped / Already Imported', readonly=True)
    result_message = fields.Text(string='Result', readonly=True)

    import_type = fields.Selection([
        ('paid_till', 'Paid Till'),
        ('sica', 'SICA/CBT'),
        ('expense', 'Expense Receipt'),
    ], default='paid_till')

    state = fields.Selection([
        ('draft', 'Draft'),
        ('done', 'Done'),
    ], default='draft')

    def _clean_header(self, value):
        return str(value or '').strip().lower().replace('.', '')

    def _to_float(self, value):
        if value in (None, '', ' '):
            return 0.0
        try:
            return float(str(value).replace(',', '').strip())
        except Exception:
            return 0.0

    def _to_date(self, value):
        if not value:
            return False

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        value = str(value).strip()

        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(value[:10], fmt).date()
            except Exception:
                pass

        return False

    def _payment_option(self, value):
        value = str(value or '').strip().lower()

        if value == 'cash':
            return 'cash'
        if 'online' in value:
            return 'online_transfer'
        if 'cheque' in value:
            return 'cheque'
        if value == 'dd':
            return 'dd'
        if 'edc' in value:
            return 'edc_machine'
        if 'sica' in value:
            return 'sica_app'

        return 'cash'

    def _state_value(self, value):
        value = str(value or '').strip().lower()
        return 'validated' if 'validated' in value else 'draft'

    def _get_membership_no(self, member_text):
        member_text = str(member_text or '').strip()

        if member_text.startswith('[') and ']' in member_text:
            return member_text.split(']')[0].replace('[', '').strip()

        parts = member_text.split()
        if parts:
            first = ''.join(filter(str.isdigit, parts[0]))
            if first:
                return first

        return ''.join(filter(str.isdigit, member_text))

    def _get(self, data, name):
        return data.get(name.lower(), False)

    def _get_any(self, data, names):
        for name in names:
            value = data.get(name.lower())
            if value not in (None, '', False):
                return value
        return False

    def _get_expense_model(self):
        action = self.env.ref(
            'member_management_system.action_expense',
            raise_if_not_found=False
        )

        if action and action.res_model:
            return self.env[action.res_model]

        raise UserError(_('Expense model not found from action_expense.'))

    def _amount_to_words(self, amount):
        try:
            currency = self.env['res.currency'].browse(20)
            return currency.amount_to_text(amount)
        except Exception:
            return str(amount) + ' Rupees'

    def action_import_paid_till(self):
        self.ensure_one()

        if not openpyxl:
            raise UserError(_('openpyxl is not installed.'))

        if not self.paid_till_file:
            raise UserError(_('Please upload Paid Till Excel file.'))

        file_data = base64.b64decode(self.paid_till_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        ws = wb.active

        updated = []
        not_found = []
        errors = []
        skipped = 0

        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index == 0:
                continue

            if not row or len(row) < 2:
                skipped += 1
                continue

            membership_no = self._get_membership_no(row[0])
            paid_till = str(row[1] or '').strip()

            if not membership_no or not paid_till:
                skipped += 1
                continue

            try:
                member = self.env['res.member'].search([
                    ('membership_no', '=', membership_no)
                ], limit=1)

                if not member:
                    not_found.append(membership_no)
                    continue

                new_paid_till = str(int(float(paid_till)))

                if str(member.paid_till or '').strip() == new_paid_till:
                    skipped += 1
                    continue

                member.write({'paid_till': new_paid_till})
                updated.append(membership_no)

            except Exception as e:
                errors.append('Row %s: %s' % (index + 1, str(e)))

        wb.close()

        return self._write_result(
            'paid_till',
            updated,
            not_found,
            errors,
            skipped,
            'Updated'
        )

    def action_import_sica(self):
        self.ensure_one()

        if not openpyxl:
            raise UserError(_('openpyxl is not installed.'))

        if not self.sica_file:
            raise UserError(_('Please upload SICA/CBT Excel file.'))

        file_data = base64.b64decode(self.sica_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_('Excel file is empty.'))

        headers = [self._clean_header(h) for h in rows[0]]

        inserted = []
        not_found = []
        errors = []
        skipped = 0

        for index, row in enumerate(rows[1:], start=2):
            try:
                data = {}
                total_count = 0

                for col_index, header in enumerate(headers):
                    value = row[col_index] if len(row) > col_index else False

                    if header == 'total amount':
                        total_count += 1
                        if total_count == 1:
                            data['sica_total'] = value
                        elif total_count == 2:
                            data['cbt_total'] = value
                    else:
                        data[header] = value

                membership_no = self._get_membership_no(self._get(data, 'member'))

                if not membership_no:
                    skipped += 1
                    continue

                member = self.env['res.member'].search([
                    ('membership_no', '=', membership_no)
                ], limit=1)

                if not member:
                    not_found.append(membership_no)
                    continue

                vals = {
                    'member_id': member.id,
                    'sica_receipt_no': str(self._get(data, 'sica receipt no') or '/').strip(),
                    'cbt_receipt_no': str(self._get(data, 'cbt receipt no') or '/').strip(),
                    'receipt_date': self._to_date(self._get(data, 'receipt date')),
                    'grade': str(self._get(data, 'grade') or '').strip().lower(),
                    'subscription_start': self._to_date(self._get(data, 'subscription from')),
                    'subscription_end': self._to_date(self._get(data, 'subscription to')),
                    'payment_option': self._payment_option(self._get(data, 'payment option')),
                    'payment_ref': str(self._get(data, 'payment reference') or '').strip(),
                    'subscription': self._to_float(self._get(data, 'subscription')),
                    'mbf': self._to_float(self._get(data, 'mbf')),
                    'fsf': self._to_float(self._get(data, 'fsf')),
                    'aifec': self._to_float(self._get(data, 'aifec')),
                    'uni_mei': self._to_float(self._get(data, 'uni mei')),
                    'dispute_amount': self._to_float(self._get(data, 'dispute amount')),
                    'dispute_category': str(self._get(data, 'dispute category') or '').strip(),
                    'mem_card_fee': self._to_float(self._get(data, 'mem card fee')),
                    'sica_donation': self._to_float(self._get(data, 'sica donation')),
                    'admission': self._to_float(self._get(data, 'admission/re-admission')),
                    'late_fee': self._to_float(self._get(data, 'late fee')),
                    'sica_misc': self._to_float(self._get(data, 'sica miscellaneous')),
                    'sica_total': self._to_float(data.get('sica_total')),
                    'cab_benefit_trust': self._to_float(self._get(data, 'cam benefit trust')),
                    'life_coverage_fund': self._to_float(self._get(data, 'life coverage fund')),
                    'cbt_misc': self._to_float(self._get(data, 'cbt miscellaneous')),
                    'cbt_donation': self._to_float(self._get(data, 'cbt donation')),
                    'cbt_total': self._to_float(data.get('cbt_total')),
                    'state': self._state_value(self._get(data, 'status')),
                }

                existing_receipt = self.env['sica.cbt.receipt'].search([
                    ('member_id', '=', member.id),
                    ('sica_receipt_no', '=', vals.get('sica_receipt_no')),
                    ('cbt_receipt_no', '=', vals.get('cbt_receipt_no')),
                ], limit=1)

                if existing_receipt:
                    skipped += 1
                    continue

                self.env['sica.cbt.receipt'].create(vals)
                inserted.append(membership_no)

            except Exception as e:
                errors.append('Row %s: %s' % (index, str(e)))

        wb.close()

        return self._write_result(
            'sica',
            inserted,
            not_found,
            errors,
            skipped,
            'Inserted'
        )

    def action_import_expense_receipt(self):
        self.ensure_one()

        if not openpyxl:
            raise UserError(_('openpyxl is not installed.'))

        if not self.expense_file:
            raise UserError(_('Please upload Expense Receipt Excel file.'))

        expense_model = self._get_expense_model()
        expense_type_model = self.env['expense.type']

        file_data = base64.b64decode(self.expense_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)

        inserted = []
        errors = []
        skipped = 0

        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            headers = [self._clean_header(h) for h in rows[0]]

            for index, row in enumerate(rows[1:], start=2):
                try:
                    data = {}

                    for col_index, header in enumerate(headers):
                        data[header] = row[col_index] if len(row) > col_index else False

                    voucher_no = str(self._get_any(data, [
                        'voucher',
                        'voucher no',
                        'voucher no.',
                        'voucher number',
                    ]) or '').strip()

                    if not voucher_no:
                        skipped += 1
                        continue

                    existing_expense = expense_model.search([
                        ('voucher_no', '=', voucher_no)
                    ], limit=1)

                    if existing_expense:
                        skipped += 1
                        continue

                    section_name = str(self._get_any(data, [
                        'section',
                        'expense type',
                        'expense_type',
                    ]) or '').strip()

                    expense_type = False

                    if section_name:
                        expense_type = expense_type_model.search([
                            ('name', '=', section_name)
                        ], limit=1)

                        if not expense_type:
                            expense_type = expense_type_model.create({
                                'name': section_name
                            })

                    amount = self._to_float(self._get_any(data, ['amount']))

                    vals = {
                        'name': str(self._get_any(data, ['to']) or '').strip(),
                        'voucher_no': voucher_no,
                        'voucher_date': self._to_date(self._get_any(data, [
                            'voucher date',
                            'date'
                        ])),
                        'expense_type_id': expense_type.id if expense_type else False,
                        'sica_cbt': str(self._get_any(data, ['sica/cbt']) or '').strip().lower(),
                        'amount': amount,
                        'amount_in_words': self._amount_to_words(amount),
                        'expense_reason': str(self._get_any(data, ['for']) or '').strip(),
                        'payment_option': self._payment_option(self._get_any(data, [
                            'payment option'
                        ])),
                        'reference_no': str(self._get_any(data, [
                            'reference no',
                            'reference no.',
                            'reference',
                        ]) or '').strip(),
                        'state': self._state_value(self._get_any(data, ['status'])),
                    }

                    expense_model.create(vals)
                    inserted.append('%s - %s' % (ws.title, voucher_no))
                    self.env.cr.commit()

                except Exception as e:
                    self.env.cr.rollback()
                    errors.append('Sheet %s Row %s: %s' % (
                        ws.title,
                        index,
                        str(e)
                    ))

        wb.close()

        return self._write_result(
            'expense',
            inserted,
            [],
            errors,
            skipped,
            'Inserted'
        )

    def _write_result(self, import_type, done_list, not_found, errors, skipped, label):
        result = ''
        result += '%s: %s\n' % (label, len(done_list))
        result += 'Not Found: %s\n' % len(not_found)
        result += 'Errors: %s\n' % len(errors)
        result += 'Skipped / Already Imported: %s\n' % skipped

        if not_found:
            result += '\n\nNot Found Members:\n'
            result += '\n'.join(not_found[:50])

        if errors:
            result += '\n\nErrors:\n'
            result += '\n'.join(errors[:20])

        self.sudo().write({
            'import_type': import_type,
            'total_records': len(done_list) + len(not_found) + len(errors) + skipped,
            'updated_count': len(done_list),
            'not_found_count': len(not_found),
            'error_count': len(errors),
            'skipped_count': skipped,
            'result_message': result,
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'member.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_reset(self):
        self.sudo().write({
            'paid_till_file': False,
            'paid_till_file_name': False,
            'sica_file': False,
            'sica_file_name': False,
            'expense_file': False,
            'expense_file_name': False,
            'state': 'draft',
            'result_message': False,
            'total_records': 0,
            'updated_count': 0,
            'not_found_count': 0,
            'error_count': 0,
            'skipped_count': 0,
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'member.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }